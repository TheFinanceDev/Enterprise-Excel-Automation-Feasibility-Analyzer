"""
Enterprise Excel Automation Feasibility Checker
=====================================================

A comprehensive tool for analyzing Excel files to determine their automation potential.
Designed for enterprise environments with security compliance in mind.

Author: Enterprise Analytics Team
Version: 1.0
Python Version: 3.7+

Required Libraries (All Standard/Common Enterprise Libraries):
- pandas: Data analysis and manipulation
- openpyxl: Excel file reading and writing
- os: Operating system interface (built-in)
- re: Regular expressions (built-in)
- datetime: Date and time handling (built-in)
- typing: Type hints (built-in Python 3.5+)
- dataclasses: Data classes (built-in Python 3.7+)

Security Notes:
- No external network connections
- No data transmission outside local environment
- Uses only standard Python libraries and common enterprise packages
- All processing done locally
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re
import os
import sys
from typing import Dict, List, Tuple, Any, Optional, Union
from dataclasses import dataclass, field
from datetime import datetime
import warnings

# Suppress pandas warnings for cleaner output
warnings.filterwarnings('ignore', category=UserWarning)
warnings.filterwarnings('ignore', category=FutureWarning)

@dataclass
class AutomationAssessment:
    """
    Comprehensive results of Excel automation feasibility analysis.
    
    This class contains all assessment results including scores, recommendations,
    detailed analysis breakdown, identified issues, and opportunities.
    """
    overall_score: float = 0.0
    feasibility_level: str = ""
    automation_recommendations: List[str] = field(default_factory=list)
    detailed_analysis: Dict[str, Any] = field(default_factory=dict)
    red_flags: List[str] = field(default_factory=list)
    opportunities: List[str] = field(default_factory=list)
    estimated_effort: str = ""
    recommended_tools: List[str] = field(default_factory=list)
    analysis_timestamp: str = ""
    file_info: Dict[str, Any] = field(default_factory=dict)

class ExcelAutomationChecker:
    """
    Main class for analyzing Excel files and determining automation feasibility.
    
    This class provides comprehensive analysis of Excel files including:
    - File structure analysis
    - Formula complexity assessment
    - Data pattern recognition
    - Business process pattern detection
    - Tool recommendations
    - Risk assessment
    """
    
    def __init__(self, verbose: bool = True):
        """
        Initialize the Excel Automation Checker.
        
        Args:
            verbose (bool): If True, prints detailed progress information
        """
        self.workbook: Optional[openpyxl.Workbook] = None
        self.file_path: Optional[str] = None
        self.analysis_results: Dict[str, Any] = {}
        self.verbose = verbose
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
        
        # Analysis configuration
        self.max_rows_to_analyze = 10000  # Limit for performance
        self.max_sheets_to_analyze = 50   # Reasonable limit for enterprise files
        
    def _log(self, message: str, level: str = "INFO") -> None:
        """
        Internal logging method for progress tracking.
        
        Args:
            message (str): Message to log
            level (str): Log level (INFO, WARNING, ERROR)
        """
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {level}: {message}")
    
    def validate_file_path(self, file_path: str) -> Tuple[bool, str]:
        """
        Validate that the provided file path is accessible and is an Excel file.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            Tuple[bool, str]: (is_valid, error_message)
        """
        try:
            # Check if file exists
            if not os.path.exists(file_path):
                return False, f"File not found: {file_path}"
            
            # Check if it's a file (not directory)
            if not os.path.isfile(file_path):
                return False, f"Path is not a file: {file_path}"
            
            # Check file extension
            _, ext = os.path.splitext(file_path.lower())
            if ext not in self.supported_extensions:
                return False, f"Unsupported file format: {ext}. Supported: {', '.join(self.supported_extensions)}"
            
            # Check file size (warn if > 100MB)
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 100:
                self._log(f"Large file detected ({file_size_mb:.1f}MB). Analysis may take longer.", "WARNING")
            
            # Check if file is accessible (not locked)
            try:
                with open(file_path, 'rb') as f:
                    f.read(1)
            except PermissionError:
                return False, "File is locked or permission denied"
            except Exception as e:
                return False, f"File access error: {str(e)}"
            
            return True, ""
            
        except Exception as e:
            return False, f"Validation error: {str(e)}"
    
    def load_excel_file(self, file_path: str) -> bool:
        """
        Load and validate Excel file for analysis.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            bool: True if file loaded successfully, False otherwise
        """
        try:
            # Validate file first
            is_valid, error_message = self.validate_file_path(file_path)
            if not is_valid:
                self._log(f"File validation failed: {error_message}", "ERROR")
                return False
            
            self.file_path = file_path
            self._log(f"Loading Excel file: {os.path.basename(file_path)}")
            
            # Load workbook with error handling for different Excel formats
            try:
                # Try to load with formulas preserved
                self.workbook = load_workbook(file_path, data_only=False, read_only=False)
                self._log("File loaded successfully with formula support")
            except Exception as e:
                self._log(f"Formula loading failed ({str(e)}), trying data-only mode", "WARNING")
                try:
                    # Fallback to data-only mode
                    self.workbook = load_workbook(file_path, data_only=True, read_only=True)
                    self._log("File loaded in data-only mode (formulas will show as values)")
                except Exception as e2:
                    self._log(f"Data-only loading failed: {str(e2)}", "ERROR")
                    return False
            
            # Validate workbook has sheets
            if not self.workbook.sheetnames:
                self._log("Error: Workbook contains no sheets", "ERROR")
                return False
            
            # Check for reasonable number of sheets
            sheet_count = len(self.workbook.sheetnames)
            if sheet_count > self.max_sheets_to_analyze:
                self._log(f"Large number of sheets detected ({sheet_count}). Analysis limited to first {self.max_sheets_to_analyze}", "WARNING")
            
            self._log(f"Successfully loaded workbook with {sheet_count} sheet(s)")
            return True
            
        except FileNotFoundError:
            self._log(f"File not found: {file_path}", "ERROR")
            return False
        except PermissionError:
            self._log("Permission denied. File may be open in Excel or locked", "ERROR")
            return False
        except Exception as e:
            self._log(f"Unexpected error loading file: {str(e)}", "ERROR")
            return False
    
    def analyze_file_structure(self) -> Dict[str, Any]:
        """
        Analyze overall file structure and organization.
        
        Returns:
            Dict[str, Any]: Comprehensive structure analysis results
        """
        if not self.workbook:
            self._log("No workbook loaded for structure analysis", "ERROR")
            return {}
        
        try:
            self._log("Analyzing file structure...")
            
            structure_analysis = {
                'total_sheets': 0,
                'sheet_names': [],
                'named_ranges': 0,
                'has_hidden_sheets': False,
                'has_very_hidden_sheets': False,
                'file_size_mb': 0.0,
                'structure_score': 0,
                'sheet_types': {
                    'visible': 0,
                    'hidden': 0,
                    'very_hidden': 0
                },
                'analysis_errors': []
            }
            
            # Basic file information
            structure_analysis['total_sheets'] = len(self.workbook.sheetnames)
            structure_analysis['sheet_names'] = self.workbook.sheetnames.copy()
            
            if self.file_path:
                structure_analysis['file_size_mb'] = round(os.path.getsize(self.file_path) / (1024 * 1024), 2)
            
            # Analyze named ranges safely
            try:
                if hasattr(self.workbook, 'defined_names'):
                    structure_analysis['named_ranges'] = len(list(self.workbook.defined_names))
                else:
                    structure_analysis['named_ranges'] = 0
            except Exception as e:
                structure_analysis['analysis_errors'].append(f"Named ranges analysis failed: {str(e)}")
                structure_analysis['named_ranges'] = 0
            
            # Analyze sheet visibility
            try:
                for sheet in self.workbook.worksheets:
                    if hasattr(sheet, 'sheet_state'):
                        if sheet.sheet_state == 'hidden':
                            structure_analysis['has_hidden_sheets'] = True
                            structure_analysis['sheet_types']['hidden'] += 1
                        elif sheet.sheet_state == 'veryHidden':
                            structure_analysis['has_very_hidden_sheets'] = True
                            structure_analysis['sheet_types']['very_hidden'] += 1
                        else:
                            structure_analysis['sheet_types']['visible'] += 1
                    else:
                        structure_analysis['sheet_types']['visible'] += 1
            except Exception as e:
                structure_analysis['analysis_errors'].append(f"Sheet visibility analysis failed: {str(e)}")
            
            # Calculate structure score (0-100)
            score = 70  # Base score
            
            # Scoring logic with error handling
            try:
                # Bonus for reasonable number of sheets (1-10 is ideal)
                if 1 <= structure_analysis['total_sheets'] <= 10:
                    score += 15
                elif 11 <= structure_analysis['total_sheets'] <= 20:
                    score += 5
                elif structure_analysis['total_sheets'] > 30:
                    score -= 20
                
                # Bonus for named ranges (shows organization)
                if structure_analysis['named_ranges'] > 0:
                    if structure_analysis['named_ranges'] <= 10:
                        score += 15  # Good organization
                    else:
                        score += 5   # Too many might be complex
                
                # Penalty for hidden sheets (adds complexity)
                if structure_analysis['has_hidden_sheets']:
                    score -= 5
                if structure_analysis['has_very_hidden_sheets']:
                    score -= 10
                
                # File size considerations
                if structure_analysis['file_size_mb'] < 5:
                    score += 5  # Small files are easier to handle
                elif structure_analysis['file_size_mb'] > 50:
                    score -= 15  # Large files are complex
                
            except Exception as e:
                structure_analysis['analysis_errors'].append(f"Scoring calculation failed: {str(e)}")
                score = 50  # Default safe score
            
            structure_analysis['structure_score'] = max(0, min(100, score))
            
            self._log(f"Structure analysis complete. Score: {structure_analysis['structure_score']}/100")
            return structure_analysis
            
        except Exception as e:
            self._log(f"Structure analysis failed: {str(e)}", "ERROR")
            return {
                'total_sheets': 0,
                'structure_score': 0,
                'analysis_errors': [f"Critical structure analysis error: {str(e)}"]
            }
    
    def analyze_sheet_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Analyze individual sheet data characteristics and structure.
        
        Args:
            sheet_name (str): Name of the sheet to analyze
            
        Returns:
            Dict[str, Any]: Detailed sheet analysis results
        """
        if not self.workbook or sheet_name not in self.workbook.sheetnames:
            self._log(f"Sheet '{sheet_name}' not found", "ERROR")
            return {}
        
        try:
            self._log(f"Analyzing sheet: {sheet_name}")
            sheet = self.workbook[sheet_name]
            
            analysis = {
                'sheet_name': sheet_name,
                'dimensions': (0, 0),
                'total_cells': 0,
                'used_cells': 0,
                'formula_cells': 0,
                'merged_cells': 0,
                'has_tables': False,
                'has_data_validation': False,
                'data_consistency_score': 0,
                'formulas': [],
                'formula_types': {},
                'headers_detected': [],
                'data_types': {},
                'blank_rows': 0,
                'blank_columns': 0,
                'formatting_complexity': 0,
                'analysis_errors': [],
                'sheet_protection': False,
                'conditional_formatting_rules': 0
            }
            
            # Get sheet dimensions safely
            try:
                max_row = sheet.max_row or 0
                max_col = sheet.max_column or 0
                analysis['dimensions'] = (max_row, max_col)
                analysis['total_cells'] = max_row * max_col
                
                # Limit analysis for very large sheets
                analysis_max_row = min(max_row, self.max_rows_to_analyze)
                if max_row > self.max_rows_to_analyze:
                    self._log(f"Large sheet detected. Analyzing first {self.max_rows_to_analyze} rows only", "WARNING")
                
            except Exception as e:
                analysis['analysis_errors'].append(f"Dimension analysis failed: {str(e)}")
                return analysis
            
            # Analyze merged cells
            try:
                analysis['merged_cells'] = len(sheet.merged_cells.ranges) if hasattr(sheet, 'merged_cells') else 0
            except Exception as e:
                analysis['analysis_errors'].append(f"Merged cells analysis failed: {str(e)}")
            
            # Check for sheet protection
            try:
                analysis['sheet_protection'] = sheet.protection.sheet if hasattr(sheet, 'protection') else False
            except Exception as e:
                analysis['analysis_errors'].append(f"Sheet protection analysis failed: {str(e)}")
            
            # Analyze cells (with performance optimization)
            formula_count = 0
            used_cell_count = 0
            data_validation_count = 0
            formula_types = {}
            
            try:
                # Iterate through cells efficiently
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=analysis_max_row, values_only=False), 1):
                    for col_idx, cell in enumerate(row, 1):
                        try:
                            if cell.value is not None:
                                used_cell_count += 1
                                
                                # Check for formulas
                                cell_value_str = str(cell.value)
                                if cell_value_str.startswith('='):
                                    formula_count += 1
                                    analysis['formulas'].append(cell_value_str[:100])  # Limit formula length for storage
                                    
                                    # Categorize formula types
                                    formula_upper = cell_value_str.upper()
                                    for func in ['SUM', 'AVERAGE', 'COUNT', 'IF', 'VLOOKUP', 'INDEX', 'MATCH', 'INDIRECT', 'OFFSET']:
                                        if func in formula_upper:
                                            formula_types[func] = formula_types.get(func, 0) + 1
                                            break
                                
                                # Check for data validation (simplified check)
                                if hasattr(cell, 'data_validation') or cell.data_type == 'f':
                                    data_validation_count += 1
                                    
                        except Exception as e:
                            # Log cell-level errors but continue processing
                            if len(analysis['analysis_errors']) < 5:  # Limit error logging
                                analysis['analysis_errors'].append(f"Cell analysis error at {row_idx},{col_idx}: {str(e)}")
                
            except Exception as e:
                analysis['analysis_errors'].append(f"Cell iteration failed: {str(e)}")
            
            analysis['used_cells'] = used_cell_count
            analysis['formula_cells'] = formula_count
            analysis['formula_types'] = formula_types
            analysis['has_data_validation'] = data_validation_count > 0
            
            # Check for Excel tables
            try:
                if hasattr(sheet, 'tables'):
                    analysis['has_tables'] = len(sheet.tables) > 0
                else:
                    analysis['has_tables'] = False
            except Exception as e:
                analysis['analysis_errors'].append(f"Table detection failed: {str(e)}")
            
            # Detect headers (first row analysis)
            try:
                if analysis['dimensions'][0] > 0:
                    first_row_values = []
                    for cell in sheet[1]:
                        if cell.value is not None:
                            first_row_values.append(str(cell.value)[:50])  # Limit header length
                        if len(first_row_values) >= 20:  # Limit number of headers to analyze
                            break
                    analysis['headers_detected'] = first_row_values
            except Exception as e:
                analysis['analysis_errors'].append(f"Header detection failed: {str(e)}")
            
            # Calculate data consistency score
            analysis['data_consistency_score'] = self._calculate_consistency_score(analysis)
            
            self._log(f"Sheet '{sheet_name}' analysis complete. Used cells: {used_cell_count}, Formulas: {formula_count}")
            return analysis
            
        except Exception as e:
            self._log(f"Critical error analyzing sheet '{sheet_name}': {str(e)}", "ERROR")
            return {
                'sheet_name': sheet_name,
                'analysis_errors': [f"Critical sheet analysis error: {str(e)}"],
                'data_consistency_score': 0
            }
    
    def analyze_formulas(self) -> Dict[str, Any]:
        """
        Analyze formula complexity and automation difficulty across all sheets.
        
        Returns:
            Dict[str, Any]: Comprehensive formula analysis results
        """
        if not self.workbook:
            self._log("No workbook loaded for formula analysis", "ERROR")
            return {}
        
        try:
            self._log("Analyzing formulas across all sheets...")
            
            all_formulas = []
            complex_functions = []
            simple_functions = []
            formula_types_summary = {}
            
            # Define complexity categories
            complex_formula_patterns = {
                'INDIRECT', 'OFFSET', 'INDEX', 'MATCH', 'VLOOKUP', 'HLOOKUP',
                'XLOOKUP', 'SUMPRODUCT', 'ARRAY', 'MMULT', 'TRANSPOSE',
                'PIVOT', 'GETPIVOTDATA', 'CUBE', 'HYPERLINK'
            }
            
            simple_formula_patterns = {
                'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'IF', 'CONCATENATE',
                'CONCAT', 'LEFT', 'RIGHT', 'MID', 'LEN', 'UPPER', 'LOWER',
                'ROUND', 'ABS', 'SQRT'
            }
            
            moderate_formula_patterns = {
                'SUMIF', 'COUNTIF', 'AVERAGEIF', 'SUMIFS', 'COUNTIFS', 'AVERAGEIFS',
                'IFERROR', 'IFNA', 'CHOOSE', 'SWITCH', 'IFS'
            }
            
            # Analyze formulas from all sheets
            sheets_analyzed = 0
            for sheet_name in self.workbook.sheetnames:
                if sheets_analyzed >= self.max_sheets_to_analyze:
                    self._log(f"Reached sheet analysis limit ({self.max_sheets_to_analyze})", "WARNING")
                    break
                
                try:
                    sheet_analysis = self.analyze_sheet_data(sheet_name)
                    sheet_formulas = sheet_analysis.get('formulas', [])
                    all_formulas.extend(sheet_formulas)
                    
                    # Update formula type summary
                    sheet_formula_types = sheet_analysis.get('formula_types', {})
                    for func, count in sheet_formula_types.items():
                        formula_types_summary[func] = formula_types_summary.get(func, 0) + count
                    
                    sheets_analyzed += 1
                    
                except Exception as e:
                    self._log(f"Error analyzing formulas in sheet '{sheet_name}': {str(e)}", "WARNING")
                    continue
            
            # Categorize formulas by complexity
            for formula in all_formulas:
                try:
                    formula_upper = formula.upper()
                    
                    # Check for complex functions
                    if any(complex_func in formula_upper for complex_func in complex_formula_patterns):
                        complex_functions.append(formula)
                    # Check for moderate functions
                    elif any(moderate_func in formula_upper for moderate_func in moderate_formula_patterns):
                        # Categorize as simple for now, but could be separate category
                        simple_functions.append(formula)
                    # Check for simple functions
                    elif any(simple_func in formula_upper for simple_func in simple_formula_patterns):
                        simple_functions.append(formula)
                    # Unknown/custom formulas are considered complex
                    else:
                        complex_functions.append(formula)
                        
                except Exception as e:
                    # If formula analysis fails, consider it complex for safety
                    complex_functions.append(formula)
            
            # Calculate metrics
            total_formulas = len(all_formulas)
            complex_count = len(complex_functions)
            simple_count = len(simple_functions)
            
            formula_analysis = {
                'total_formulas': total_formulas,
                'complex_formulas': complex_count,
                'simple_formulas': simple_count,
                'formula_complexity_ratio': complex_count / max(1, total_formulas),
                'formula_types_summary': formula_types_summary,
                'most_complex_formulas': complex_functions[:5],  # Top 5 most complex
                'automation_difficulty_score': 0,
                'sheets_analyzed': sheets_analyzed,
                'analysis_errors': []
            }
            
            # Calculate automation difficulty based on formulas
            try:
                if total_formulas == 0:
                    difficulty_score = 20  # No formulas = easy to automate
                else:
                    complexity_ratio = formula_analysis['formula_complexity_ratio']
                    if complexity_ratio > 0.6:
                        difficulty_score = 85  # Very high complexity
                    elif complexity_ratio > 0.4:
                        difficulty_score = 70  # High complexity
                    elif complexity_ratio > 0.2:
                        difficulty_score = 50  # Medium complexity
                    else:
                        difficulty_score = 25  # Low complexity
                
                # Adjust based on total formula count
                if total_formulas > 1000:
                    difficulty_score += 10  # Many formulas add complexity
                elif total_formulas > 100:
                    difficulty_score += 5
                
                formula_analysis['automation_difficulty_score'] = min(100, difficulty_score)
                
            except Exception as e:
                formula_analysis['analysis_errors'].append(f"Difficulty scoring failed: {str(e)}")
                formula_analysis['automation_difficulty_score'] = 50  # Default moderate difficulty
            
            self._log(f"Formula analysis complete. Total: {total_formulas}, Complex: {complex_count}, Difficulty: {formula_analysis['automation_difficulty_score']}/100")
            return formula_analysis
            
        except Exception as e:
            self._log(f"Critical error in formula analysis: {str(e)}", "ERROR")
            return {
                'total_formulas': 0,
                'automation_difficulty_score': 100,  # Assume high difficulty on error
                'analysis_errors': [f"Critical formula analysis error: {str(e)}"]
            }
    
    def detect_automation_patterns(self) -> Dict[str, Any]:
        """
        Detect patterns that indicate good automation potential.
        
        Returns:
            Dict[str, Any]: Pattern analysis results
        """
        if not self.workbook:
            self._log("No workbook loaded for pattern analysis", "ERROR")
            return {}
        
        try:
            self._log("Detecting automation patterns...")
            
            patterns = {
                'repeated_structures': False,
                'template_sheets': [],
                'data_entry_sheets': [],
                'calculation_sheets': [],
                'reporting_sheets': [],
                'summary_sheets': [],
                'time_based_patterns': False,
                'consolidation_patterns': False,
                'pattern_score': 0,
                'detected_patterns': [],
                'business_process_indicators': [],
                'analysis_errors': []
            }
            
            sheet_names = self.workbook.sheetnames
            
            # Time-based pattern detection (more comprehensive)
            time_patterns = {
                'months': ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                          'jul', 'aug', 'sep', 'oct', 'nov', 'dec',
                          'january', 'february', 'march', 'april', 'june',
                          'july', 'august', 'september', 'october', 'november', 'december'],
                'quarters': ['q1', 'q2', 'q3', 'q4', 'quarter'],
                'years': ['2020', '2021', '2022', '2023', '2024', '2025', '2026'],
                'periods': ['week', 'daily', 'monthly', 'annual', 'yearly']
            }
            
            time_based_sheets = []
            for name in sheet_names:
                name_lower = name.lower()
                for pattern_type, pattern_list in time_patterns.items():
                    if any(pattern in name_lower for pattern in pattern_list):
                        time_based_sheets.append(name)
                        break
            
            if len(time_based_sheets) >= 2:
                patterns['time_based_patterns'] = True
                patterns['pattern_score'] += 30
                patterns['detected_patterns'].append("Time-based sheet structure detected")
                patterns['business_process_indicators'].append("Periodic reporting process")
            
            # Repeated structure detection (improved algorithm)
            try:
                if len(sheet_names) >= 3:
                    # Check for naming patterns
                    pattern_groups = {}
                    for name in sheet_names:
                        # Extract potential base name (remove numbers, common suffixes)
                        base_name = re.sub(r'[0-9]+|_\d+|\s+\d+', '', name).strip()
                        base_name = re.sub(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec).*', '', base_name.lower()).strip()
                        
                        if len(base_name) >= 2:  # Minimum meaningful base name
                            pattern_groups[base_name] = pattern_groups.get(base_name, 0) + 1
                    
                    # Find groups with multiple sheets
                    repeated_patterns = {k: v for k, v in pattern_groups.items() if v >= 2}
                    
                    if repeated_patterns:
                        patterns['repeated_structures'] = True
                        patterns['pattern_score'] += 25
                        patterns['detected_patterns'].append("Repeated sheet naming patterns detected")
                        patterns['business_process_indicators'].append("Template-based workflow")
                        
            except Exception as e:
                patterns['analysis_errors'].append(f"Repeated structure detection failed: {str(e)}")
            
            # Categorize sheets by function (improved categorization)
            categorization_rules = {
                'data_entry': ['input', 'entry', 'data', 'raw', 'import', 'source', 'form'],
                'calculation': ['calc', 'calculation', 'compute', 'process', 'analysis', 'logic'],
                'reporting': ['report', 'summary', 'dashboard', 'output', 'results', 'final'],
                'template': ['template', 'master', 'base', 'model', 'standard'],
                'summary': ['summary', 'total', 'consolidated', 'overview', 'aggregate']
            }
            
            for sheet_name in sheet_names:
                name_lower = sheet_name.lower()
                categorized = False
                
                for category, keywords in categorization_rules.items():
                    if any(keyword in name_lower for keyword in keywords):
                        patterns[f'{category}_sheets'].append(sheet_name)
                        categorized = True
                        break
                
                if not categorized:
                    # Try to categorize based on sheet content hints
                    # This is a simplified approach - in practice, you'd analyze sheet content
                    if any(word in name_lower for word in ['sheet', 'tab', 'page']):
                        patterns['data_entry_sheets'].append(sheet_name)
            
            # Detect consolidation patterns
            data_sources = len(patterns['data_entry_sheets'])
            reporting_outputs = len(patterns['reporting_sheets']) + len(patterns['summary_sheets'])
            
            if data_sources >= 2 and reporting_outputs >= 1:
                patterns['consolidation_patterns'] = True
                patterns['pattern_score'] += 20
                patterns['detected_patterns'].append("Data consolidation workflow detected")
                patterns['business_process_indicators'].append("Multi-source data aggregation")
            
            # Additional pattern detection
            
            # Template pattern (master + variations)
            if patterns['template_sheets'] and (len(sheet_names) - len(patterns['template_sheets'])) >= 2:
                patterns['pattern_score'] += 15
                patterns['detected_patterns'].append("Master template with variations")
                patterns['business_process_indicators'].append("Standardized reporting process")
            
            # Complex business process pattern
            if (patterns['data_entry_sheets'] and 
                patterns['calculation_sheets'] and 
                patterns['reporting_sheets']):
                patterns['pattern_score'] += 25
                patterns['detected_patterns'].append("Complete business process workflow")
                patterns['business_process_indicators'].append("End-to-end data processing pipeline")
            
            # Workflow complexity bonus
            if len(patterns['detected_patterns']) >= 3:
                patterns['pattern_score'] += 10
                patterns['business_process_indicators'].append("Multi-pattern business process")
            
            self._log(f"Pattern analysis complete. Score: {patterns['pattern_score']}/100")
            return patterns
            
        except Exception as e:
            self._log(f"Critical error in pattern analysis: {str(e)}", "ERROR")
            return {
                'pattern_score': 0,
                'analysis_errors': [f"Critical pattern analysis error: {str(e)}"]
            }
    
    def identify_red_flags(self) -> List[str]:
        """
        Identify factors that make automation difficult or risky.
        
        Returns:
            List[str]: List of identified red flags and issues
        """
        if not self.workbook:
            self._log("No workbook loaded for red flag analysis", "ERROR")
            return ["No workbook loaded"]
        
        try:
            self._log("Identifying automation red flags...")
            red_flags = []
            
            # Get overall file metrics
            structure_analysis = self.analysis_results.get('structure', {})
            formula_analysis = self.analysis_results.get('formulas', {})
            
            # File-level red flags
            file_size_mb = structure_analysis.get('file_size_mb', 0)
            if file_size_mb > 100:
                red_flags.append(f"Very large file size ({file_size_mb:.1f}MB) - may cause performance issues")
            elif file_size_mb > 50:
                red_flags.append(f"Large file size ({file_size_mb:.1f}MB) - requires optimization considerations")
            
            # Sheet count red flags
            total_sheets = structure_analysis.get('total_sheets', 0)
            if total_sheets > 30:
                red_flags.append(f"Excessive number of sheets ({total_sheets}) - increases automation complexity")
            elif total_sheets > 20:
                red_flags.append(f"High number of sheets ({total_sheets}) - may require phased automation approach")
            
            # Formula complexity red flags
            complex_formula_ratio = formula_analysis.get('formula_complexity_ratio', 0)
            total_formulas = formula_analysis.get('total_formulas', 0)
            
            if complex_formula_ratio > 0.7:
                red_flags.append("Very high complex formula ratio (>70%) - significant logic replication required")
            elif complex_formula_ratio > 0.5:
                red_flags.append("High complex formula ratio (>50%) - moderate logic replication required")
            
            if total_formulas > 2000:
                red_flags.append(f"Very high formula count ({total_formulas}) - extensive logic to replicate")
            elif total_formulas > 1000:
                red_flags.append(f"High formula count ({total_formulas}) - substantial automation effort required")
            
            # Analyze individual sheets for red flags
            total_merged_cells = 0
            sheets_with_protection = 0
            sheets_with_many_formulas = 0
            
            for sheet_name in self.workbook.sheetnames[:self.max_sheets_to_analyze]:
                try:
                    sheet_analysis = self.analyze_sheet_data(sheet_name)
                    
                    # Merged cells analysis
                    merged_count = sheet_analysis.get('merged_cells', 0)
                    total_merged_cells += merged_count
                    
                    # Sheet protection
                    if sheet_analysis.get('sheet_protection', False):
                        sheets_with_protection += 1
                    
                    # Formula density per sheet
                    used_cells = sheet_analysis.get('used_cells', 1)
                    formula_cells = sheet_analysis.get('formula_cells', 0)
                    if used_cells > 100 and (formula_cells / used_cells) > 0.7:
                        sheets_with_many_formulas += 1
                    
                    # Sheet-specific red flags
                    if merged_count > 50:
                        red_flags.append(f"Sheet '{sheet_name}' has excessive merged cells ({merged_count})")
                    
                    # Check for very sparse data
                    total_cells = sheet_analysis.get('total_cells', 1)
                    if used_cells > 0 and total_cells > 1000:
                        density = used_cells / total_cells
                        if density < 0.05:  # Less than 5% of cells used
                            red_flags.append(f"Sheet '{sheet_name}' has very sparse data (low density)")
                    
                except Exception as e:
                    red_flags.append(f"Analysis error in sheet '{sheet_name}': {str(e)}")
                    continue
            
            # Aggregate red flags
            if total_merged_cells > 20:
                red_flags.append(f"High total merged cells across file ({total_merged_cells}) - major automation blocker")
            elif total_merged_cells > 10:
                red_flags.append(f"Moderate merged cell usage ({total_merged_cells}) - complicates automation")
            
            if sheets_with_protection > 0:
                red_flags.append(f"Protected sheets detected ({sheets_with_protection}) - may restrict automation access")
            
            if sheets_with_many_formulas > (total_sheets * 0.5):
                red_flags.append("Many sheets are formula-heavy - complex logic replication required")
            
            # Hidden sheets red flag
            if structure_analysis.get('has_hidden_sheets', False):
                red_flags.append("Hidden sheets detected - may contain critical logic or data")
            
            if structure_analysis.get('has_very_hidden_sheets', False):
                red_flags.append("Very hidden sheets detected - likely contains sensitive or complex logic")
            
            # Check for VBA/Macros (enhanced detection)
            try:
                # Multiple ways to detect macros
                has_macros = False
                
                # Method 1: Check file extension
                if self.file_path and self.file_path.lower().endswith('.xlsm'):
                    has_macros = True
                
                # Method 2: Check for VBA archive
                if hasattr(self.workbook, 'vba_archive') and self.workbook.vba_archive:
                    has_macros = True
                
                # Method 3: Look for macro-related function calls in formulas
                formula_types = formula_analysis.get('formula_types_summary', {})
                macro_indicators = ['CALL', 'RUN', 'PERSONAL', 'APPLICATION']
                if any(indicator in formula_types for indicator in macro_indicators):
                    has_macros = True
                
                if has_macros:
                    red_flags.append("VBA macros detected - requires specialized automation approach")
                    
            except Exception as e:
                red_flags.append(f"Macro detection failed: {str(e)}")
            
            # Data quality red flags
            try:
                consistency_scores = []
                for sheet_name in self.workbook.sheetnames[:10]:  # Check first 10 sheets
                    sheet_analysis = self.analyze_sheet_data(sheet_name)
                    consistency_scores.append(sheet_analysis.get('data_consistency_score', 50))
                
                if consistency_scores:
                    avg_consistency = sum(consistency_scores) / len(consistency_scores)
                    if avg_consistency < 40:
                        red_flags.append("Low data consistency across sheets - requires cleanup before automation")
                    elif avg_consistency < 60:
                        red_flags.append("Moderate data consistency issues - some cleanup recommended")
                        
            except Exception as e:
                red_flags.append(f"Data consistency check failed: {str(e)}")
            
            self._log(f"Red flag analysis complete. Found {len(red_flags)} potential issues")
            return red_flags
            
        except Exception as e:
            self._log(f"Critical error in red flag analysis: {str(e)}", "ERROR")
            return [f"Critical red flag analysis error: {str(e)}"]
    
    def identify_opportunities(self) -> List[str]:
        """
        Identify specific automation opportunities and benefits.
        
        Returns:
            List[str]: List of automation opportunities
        """
        try:
            self._log("Identifying automation opportunities...")
            opportunities = []
            
            # Get analysis results
            structure_analysis = self.analysis_results.get('structure', {})
            formula_analysis = self.analysis_results.get('formulas', {})
            pattern_analysis = self.analysis_results.get('patterns', {})
            
            # Pattern-based opportunities
            if pattern_analysis.get('time_based_patterns', False):
                opportunities.append("Monthly/quarterly reporting can be fully automated with templates")
            
            if pattern_analysis.get('consolidation_patterns', False):
                opportunities.append("Data consolidation process has high ROI automation potential")
            
            if pattern_analysis.get('repeated_structures', False):
                opportunities.append("Repeated sheet structures can use template-based automation")
            
            # Structure-based opportunities
            if structure_analysis.get('named_ranges', 0) > 0:
                opportunities.append("Named ranges indicate well-structured data - easier automation implementation")
            
            if structure_analysis.get('total_sheets', 0) <= 10:
                opportunities.append("Manageable number of sheets - straightforward automation scope")
            
            # Formula-based opportunities
            simple_formulas = formula_analysis.get('simple_formulas', 0)
            total_formulas = formula_analysis.get('total_formulas', 1)
            
            if simple_formulas > (total_formulas * 0.7):
                opportunities.append("Majority of formulas are simple - easy to replicate in automation")
            
            if total_formulas < 100:
                opportunities.append("Low formula complexity - minimal logic replication required")
            
            # Business process opportunities
            business_indicators = pattern_analysis.get('business_process_indicators', [])
            if 'End-to-end data processing pipeline' in business_indicators:
                opportunities.append("Complete workflow automation possible - high impact potential")
            
            if 'Multi-source data aggregation' in business_indicators:
                opportunities.append("Data integration automation can eliminate manual consolidation")
            
            # File characteristics opportunities
            file_size = structure_analysis.get('file_size_mb', 0)
            if file_size < 10:
                opportunities.append("Small file size enables cloud-based automation solutions")
            
            # Sheet type opportunities
            if pattern_analysis.get('template_sheets'):
                opportunities.append("Master templates detected - can standardize and automate variations")
            
            if len(pattern_analysis.get('data_entry_sheets', [])) > 0:
                opportunities.append("Data entry processes can be automated with forms or APIs")
            
            if len(pattern_analysis.get('reporting_sheets', [])) > 0:
                opportunities.append("Report generation can be automated with scheduled processes")
            
            # Add ROI indicators
            if len(opportunities) >= 5:
                opportunities.append("Multiple automation opportunities identified - high ROI potential")
            elif len(opportunities) >= 3:
                opportunities.append("Several automation opportunities - good ROI potential")
            
            self._log(f"Opportunity analysis complete. Found {len(opportunities)} opportunities")
            return opportunities
            
        except Exception as e:
            self._log(f"Error identifying opportunities: {str(e)}", "ERROR")
            return ["Opportunity analysis failed - manual review recommended"]
    
    def recommend_automation_tools(self, analysis_summary: Dict[str, Any]) -> List[str]:
        """
        Recommend appropriate automation tools based on comprehensive analysis.
        
        Args:
            analysis_summary (Dict[str, Any]): Summary of all analysis results
            
        Returns:
            List[str]: Ordered list of recommended automation tools with rationale
        """
        try:
            self._log("Generating automation tool recommendations...")
            recommendations = []
            
            # Extract key metrics
            overall_score = analysis_summary.get('overall_score', 0)
            formula_complexity = analysis_summary.get('formula_complexity_ratio', 0)
            file_size = analysis_summary.get('file_size_mb', 0)
            has_patterns = analysis_summary.get('has_automation_patterns', False)
            total_formulas = analysis_summary.get('total_formulas', 0)
            has_macros = any('macro' in flag.lower() or 'vba' in flag.lower() 
                           for flag in analysis_summary.get('red_flags', []))
            
            # Python-based solutions (best for clean, structured data)
            if overall_score >= 70 and formula_complexity < 0.4 and not has_macros:
                recommendations.extend([
                    "✅ Python + pandas + openpyxl - Ideal for data processing, calculations, and report generation",
                    "✅ Python + xlwings - Excel integration with Python logic, maintains Excel interface"
                ])
            elif overall_score >= 50 and formula_complexity < 0.6:
                recommendations.append("⚠️ Python + pandas - Possible with formula logic replication effort")
            
            # Microsoft Power Platform (good for enterprise integration)
            if overall_score >= 60 and file_size < 30 and not has_macros:
                recommendations.extend([
                    "✅ Microsoft Power Automate - Excellent for workflow automation and Office 365 integration",
                    "✅ Power BI + Power Query - Perfect for reporting automation and data transformation"
                ])
            elif overall_score >= 40:
                recommendations.append("⚠️ Power Platform - Possible with data restructuring")
            
            # VBA/Excel-based solutions (for complex Excel-native logic)
            if has_macros or formula_complexity > 0.5 or total_formulas > 500:
                recommendations.extend([
                    "✅ Enhanced VBA/Excel Macros - Build upon existing logic, add automation triggers",
                    "✅ Excel + VBA + Python integration - Hybrid approach leveraging both platforms"
                ])
            elif overall_score >= 60:
                recommendations.append("✅ Excel VBA - Native Excel automation for formula-heavy processes")
            
            # RPA solutions (for complex, hard-to-restructure processes)
            if overall_score < 50 or len(analysis_summary.get('red_flags', [])) > 3:
                recommendations.extend([
                    "⚠️ RPA Tools (UiPath, Automation Anywhere) - For processes that can't be restructured",
                    "⚠️ Desktop automation - When existing file structure must be preserved"
                ])
            
            # Cloud solutions (for collaborative environments)
            if has_patterns and file_size < 20 and overall_score >= 60:
                recommendations.extend([
                    "✅ Google Sheets + Apps Script - Cloud-based collaborative automation",
                    "✅ Office 365 + SharePoint + Power Automate - Enterprise cloud automation suite"
                ])
            
            # Specialized solutions based on specific patterns
            pattern_analysis = self.analysis_results.get('patterns', {})
            
            if pattern_analysis.get('time_based_patterns', False):
                recommendations.append("✅ Scheduled automation scripts - Perfect for periodic reporting")
            
            if pattern_analysis.get('consolidation_patterns', False):
                recommendations.append("✅ ETL tools (SSIS, Alteryx) - Specialized for data consolidation workflows")
            
            # No-code/low-code options
            if overall_score >= 50 and formula_complexity < 0.3:
                recommendations.append("✅ No-code platforms (Zapier, Microsoft Flow) - Quick implementation option")
            
            # Remove duplicates while preserving order
            seen = set()
            unique_recommendations = []
            for rec in recommendations:
                if rec not in seen:
                    seen.add(rec)
                    unique_recommendations.append(rec)
            
            # Limit to top 8 recommendations
            final_recommendations = unique_recommendations[:8]
            
            self._log(f"Generated {len(final_recommendations)} tool recommendations")
            return final_recommendations
            
        except Exception as e:
            self._log(f"Error generating tool recommendations: {str(e)}", "ERROR")
            return ["Tool recommendation analysis failed - manual assessment required"]
    
    def _calculate_consistency_score(self, sheet_analysis: Dict[str, Any]) -> float:
        """
        Calculate data consistency score for a sheet.
        
        Args:
            sheet_analysis (Dict[str, Any]): Sheet analysis results
            
        Returns:
            float: Consistency score (0-100)
        """
        try:
            score = 70  # Base score
            
            # Bonus factors
            if sheet_analysis.get('has_tables', False):
                score += 20  # Excel tables indicate structure
            
            if sheet_analysis.get('has_data_validation', False):
                score += 15  # Data validation shows controlled input
            
            if len(sheet_analysis.get('headers_detected', [])) > 0:
                score += 10  # Headers indicate structure
            
            # Penalty factors
            used_cells = max(1, sheet_analysis.get('used_cells', 1))
            
            # Merged cells penalty
            merged_cells = sheet_analysis.get('merged_cells', 0)
            if merged_cells > 0:
                merged_ratio = merged_cells / used_cells
                if merged_ratio > 0.2:
                    score -= 30  # High merged cell ratio
                elif merged_ratio > 0.1:
                    score -= 15  # Moderate merged cell ratio
                else:
                    score -= 5   # Low merged cell ratio
            
            # Data density penalty
            total_cells = max(1, sheet_analysis.get('total_cells', 1))
            if used_cells > 0 and total_cells > 100:
                density = used_cells / total_cells
                if density < 0.05:
                    score -= 20  # Very sparse
                elif density < 0.1:
                    score -= 10  # Moderately sparse
            
            # Formula ratio considerations
            formula_cells = sheet_analysis.get('formula_cells', 0)
            if used_cells > 10:  # Only for sheets with meaningful data
                formula_ratio = formula_cells / used_cells
                if formula_ratio > 0.8:
                    score -= 10  # Too formula-heavy might indicate complex logic
                elif 0.1 <= formula_ratio <= 0.5:
                    score += 5   # Good balance of data and calculations
            
            return max(0, min(100, score))
            
        except Exception as e:
            self._log(f"Error calculating consistency score: {str(e)}", "WARNING")
            return 50.0  # Default moderate score on error
    
    def generate_comprehensive_report(self) -> AutomationAssessment:
        """
        Generate final comprehensive automation feasibility assessment.
        
        Returns:
            AutomationAssessment: Complete analysis results and recommendations
        """
        if not self.workbook:
            raise ValueError("No Excel file loaded. Please load a file first using load_excel_file()")
        
        try:
            self._log("Generating comprehensive automation assessment...")
            
            # Perform all analyses with error handling
            try:
                structure_analysis = self.analyze_file_structure()
                self.analysis_results['structure'] = structure_analysis
            except Exception as e:
                self._log(f"Structure analysis failed: {str(e)}", "ERROR")
                structure_analysis = {'structure_score': 0, 'analysis_errors': [str(e)]}
                self.analysis_results['structure'] = structure_analysis
            
            try:
                formula_analysis = self.analyze_formulas()
                self.analysis_results['formulas'] = formula_analysis
            except Exception as e:
                self._log(f"Formula analysis failed: {str(e)}", "ERROR")
                formula_analysis = {'automation_difficulty_score': 100, 'analysis_errors': [str(e)]}
                self.analysis_results['formulas'] = formula_analysis
            
            try:
                pattern_analysis = self.detect_automation_patterns()
                self.analysis_results['patterns'] = pattern_analysis
            except Exception as e:
                self._log(f"Pattern analysis failed: {str(e)}", "ERROR")
                pattern_analysis = {'pattern_score': 0, 'analysis_errors': [str(e)]}
                self.analysis_results['patterns'] = pattern_analysis
            
            try:
                red_flags = self.identify_red_flags()
                self.analysis_results['red_flags'] = red_flags
            except Exception as e:
                self._log(f"Red flag analysis failed: {str(e)}", "ERROR")
                red_flags = [f"Red flag analysis error: {str(e)}"]
                self.analysis_results['red_flags'] = red_flags
            
            try:
                opportunities = self.identify_opportunities()
            except Exception as e:
                self._log(f"Opportunity analysis failed: {str(e)}", "ERROR")
                opportunities = [f"Opportunity analysis error: {str(e)}"]
            
            # Calculate overall score with error handling
            try:
                structure_score = structure_analysis.get('structure_score', 0)
                formula_difficulty = 100 - formula_analysis.get('automation_difficulty_score', 50)
                pattern_score = pattern_analysis.get('pattern_score', 0)
                
                # Weighted scoring
                overall_score = (
                    structure_score * 0.3 +
                    formula_difficulty * 0.4 +
                    pattern_score * 0.3
                )
                
                # Apply red flag penalties
                red_flag_penalty = min(len(red_flags) * 8, 40)  # Max 40 point penalty
                overall_score = max(0, overall_score - red_flag_penalty)
                
            except Exception as e:
                self._log(f"Score calculation failed: {str(e)}", "ERROR")
                overall_score = 25.0  # Conservative default
            
            # Determine feasibility level and effort estimation
            try:
                if overall_score >= 80:
                    feasibility_level = "HIGH - Excellent automation candidate"
                    estimated_effort = "2-4 weeks development + 1 week testing"
                elif overall_score >= 65:
                    feasibility_level = "MEDIUM-HIGH - Very good automation potential"
                    estimated_effort = "1-2 months development + 2 weeks testing"
                elif overall_score >= 50:
                    feasibility_level = "MEDIUM - Good automation potential with preparation"
                    estimated_effort = "2-3 months development + 3 weeks testing"
                elif overall_score >= 35:
                    feasibility_level = "LOW-MEDIUM - Possible but requires significant restructuring"
                    estimated_effort = "3-4 months development + 1 month testing"
                else:
                    feasibility_level = "LOW - Not recommended for automation in current state"
                    estimated_effort = "6+ months or complete redesign recommended"
                    
            except Exception as e:
                self._log(f"Feasibility level calculation failed: {str(e)}", "ERROR")
                feasibility_level = "UNKNOWN - Analysis incomplete"
                estimated_effort = "Manual assessment required"
            
            # Generate tool recommendations
            try:
                analysis_summary = {
                    'overall_score': overall_score,
                    'formula_complexity_ratio': formula_analysis.get('formula_complexity_ratio', 0),
                    'file_size_mb': structure_analysis.get('file_size_mb', 0),
                    'has_automation_patterns': pattern_analysis.get('pattern_score', 0) > 20,
                    'red_flags': red_flags,
                    'total_formulas': formula_analysis.get('total_formulas', 0)
                }
                
                recommended_tools = self.recommend_automation_tools(analysis_summary)
                
            except Exception as e:
                self._log(f"Tool recommendation failed: {str(e)}", "ERROR")
                recommended_tools = ["Tool recommendation failed - manual assessment required"]
            
            # Create detailed analysis summary
            try:
                detailed_analysis = {
                    'file_info': {
                        'file_name': os.path.basename(self.file_path) if self.file_path else "Unknown",
                        'file_path': self.file_path or "Unknown",
                        'sheets': structure_analysis.get('total_sheets', 0),
                        'file_size_mb': round(structure_analysis.get('file_size_mb', 0), 2),
                        'named_ranges': structure_analysis.get('named_ranges', 0),
                        'has_hidden_sheets': structure_analysis.get('has_hidden_sheets', False)
                    },
                    'formula_analysis': {
                        'total_formulas': formula_analysis.get('total_formulas', 0),
                        'complex_formulas': formula_analysis.get('complex_formulas', 0),
                        'simple_formulas': formula_analysis.get('simple_formulas', 0),
                        'complexity_ratio': round(formula_analysis.get('formula_complexity_ratio', 0), 3),
                        'formula_types': formula_analysis.get('formula_types_summary', {}),
                        'difficulty_score': formula_analysis.get('automation_difficulty_score', 0)
                    },
                    'pattern_analysis': {
                        'time_based_patterns': pattern_analysis.get('time_based_patterns', False),
                        'repeated_structures': pattern_analysis.get('repeated_structures', False),
                        'consolidation_patterns': pattern_analysis.get('consolidation_patterns', False),
                        'detected_patterns': pattern_analysis.get('detected_patterns', []),
                        'business_indicators': pattern_analysis.get('business_process_indicators', [])
                    },
                    'scores': {
                        'structure_score': round(structure_score, 1),
                        'formula_difficulty_score': round(100 - formula_analysis.get('automation_difficulty_score', 50), 1),
                        'pattern_score': round(pattern_score, 1),
                        'overall_score': round(overall_score, 1)
                    },
                    'quality_metrics': {
                        'red_flags_count': len(red_flags),
                        'opportunities_count': len(opportunities),
                        'analysis_completeness': self._calculate_analysis_completeness()
                    }
                }
            except Exception as e:
                self._log(f"Detailed analysis summary creation failed: {str(e)}", "ERROR")
                detailed_analysis = {'error': f"Summary creation failed: {str(e)}"}
            
            # Create final assessment
            assessment = AutomationAssessment(
                overall_score=round(overall_score, 1),
                feasibility_level=feasibility_level,
                automation_recommendations=recommended_tools,
                detailed_analysis=detailed_analysis,
                red_flags=red_flags,
                opportunities=opportunities,
                estimated_effort=estimated_effort,
                recommended_tools=recommended_tools,
                analysis_timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                file_info={
                    'name': os.path.basename(self.file_path) if self.file_path else "Unknown",
                    'size_mb': structure_analysis.get('file_size_mb', 0),
                    'sheets': structure_analysis.get('total_sheets', 0)
                }
            )
            
            self._log("Comprehensive assessment completed successfully")
            return assessment
            
        except Exception as e:
            self._log(f"Critical error generating comprehensive report: {str(e)}", "ERROR")
            # Return a safe fallback assessment
            return AutomationAssessment(
                overall_score=0.0,
                feasibility_level="ERROR - Analysis failed",
                automation_recommendations=["Manual assessment required due to analysis failure"],
                detailed_analysis={'error': str(e)},
                red_flags=[f"Critical analysis error: {str(e)}"],
                opportunities=[],
                estimated_effort="Unknown - manual review required",
                recommended_tools=["Manual assessment required"],
                analysis_timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                file_info={'name': 'Error', 'size_mb': 0, 'sheets': 0}
            )
    
    def _calculate_analysis_completeness(self) -> float:
        """
        Calculate how complete the analysis was (useful for quality assessment).
        
        Returns:
            float: Completeness percentage (0-100)
        """
        try:
            total_components = 4  # structure, formulas, patterns, red_flags
            completed_components = 0
            
            if 'structure' in self.analysis_results and self.analysis_results['structure']:
                completed_components += 1
            if 'formulas' in self.analysis_results and self.analysis_results['formulas']:
                completed_components += 1
            if 'patterns' in self.analysis_results and self.analysis_results['patterns']:
                completed_components += 1
            if 'red_flags' in self.analysis_results:
                completed_components += 1
            
            return (completed_components / total_components) * 100
            
        except Exception:
            return 50.0  # Default moderate completeness on error
    
    def export_report_to_text(self, assessment: AutomationAssessment, output_path: Optional[str] = None) -> str:
        """
        Export the assessment report to a formatted text file.
        
        Args:
            assessment (AutomationAssessment): The assessment results to export
            output_path (Optional[str]): Path for output file. If None, uses default naming
            
        Returns:
            str: Path to the exported report file
        """
        try:
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(self.file_path or "unknown"))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"Excel_Automation_Report_{base_name}_{timestamp}.txt"
            
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write("RECOMMENDED AUTOMATION TOOLS\n")
                f.write("-" * 40 + "\n")
                for i, tool in enumerate(assessment.recommended_tools, 1):
                    f.write(f"{i}. {tool}\n")
                f.write("\n")
                
                # Detailed analysis section
                if 'scores' in assessment.detailed_analysis:
                    f.write("DETAILED ANALYSIS SCORES\n")
                    f.write("-" * 40 + "\n")
                    scores = assessment.detailed_analysis['scores']
                    f.write(f"Structure Score: {scores.get('structure_score', 0)}/100\n")
                    f.write(f"Formula Difficulty Score: {scores.get('formula_difficulty_score', 0)}/100\n")
                    f.write(f"Pattern Recognition Score: {scores.get('pattern_score', 0)}/100\n")
                    f.write(f"Overall Score: {scores.get('overall_score', 0)}/100\n\n")
                
                # Technical details
                if 'formula_analysis' in assessment.detailed_analysis:
                    f.write("TECHNICAL ANALYSIS\n")
                    f.write("-" * 40 + "\n")
                    formula_info = assessment.detailed_analysis['formula_analysis']
                    f.write(f"Total Formulas: {formula_info.get('total_formulas', 0)}\n")
                    f.write(f"Complex Formulas: {formula_info.get('complex_formulas', 0)}\n")
                    f.write(f"Simple Formulas: {formula_info.get('simple_formulas', 0)}\n")
                    f.write(f"Complexity Ratio: {formula_info.get('complexity_ratio', 0):.1%}\n\n")
                
                f.write("END OF REPORT\n")
                f.write("=" * 80 + "\n")
            
            self._log(f"Report exported to: {output_path}")
            return output_path
            
        except Exception as e:
            self._log(f"Failed to export report: {str(e)}", "ERROR")
            return ""

class ExcelAutomationBot:
    """
    Main bot interface for Excel automation feasibility checking.
    
    This class provides a user-friendly interface for the automation checker
    and handles all user interactions, file processing, and report generation.
    """
    
    def __init__(self, verbose: bool = True):
        """
        Initialize the Excel Automation Bot.
        
        Args:
            verbose (bool): If True, shows detailed progress information
        """
        self.checker = ExcelAutomationChecker(verbose=verbose)
        self.last_assessment: Optional[AutomationAssessment] = None
        
    def check_requirements(self) -> bool:
        """
        Check if all required libraries are available.
        
        Returns:
            bool: True if all requirements are met
        """
        try:
            required_modules = {
                'pandas': 'Data analysis library',
                'openpyxl': 'Excel file processing library',
                'os': 'Operating system interface (built-in)',
                'datetime': 'Date/time handling (built-in)',
                're': 'Regular expressions (built-in)'
            }
            
            missing_modules = []
            
            for module, description in required_modules.items():
                try:
                    if module == 'pandas':
                        import pandas
                    elif module == 'openpyxl':
                        import openpyxl
                    elif module == 'os':
                        import os
                    elif module == 'datetime':
                        import datetime
                    elif module == 're':
                        import re
                except ImportError:
                    missing_modules.append(f"{module} - {description}")
            
            if missing_modules:
                print("❌ Missing required libraries:")
                for module in missing_modules:
                    print(f"   • {module}")
                print("\nPlease install missing libraries using:")
                print("pip install pandas openpyxl")
                return False
            
            print("✅ All required libraries are available")
            return True
            
        except Exception as e:
            print(f"❌ Error checking requirements: {str(e)}")
            return False
    
    def analyze_file(self, file_path: str) -> Optional[AutomationAssessment]:
        """
        Analyze an Excel file for automation feasibility.
        
        Args:
            file_path (str): Path to the Excel file to analyze
            
        Returns:
            Optional[AutomationAssessment]: Assessment results or None if failed
        """
        try:
            print(f"\n🔍 Starting analysis of: {os.path.basename(file_path)}")
            print("=" * 80)
            
            # Load the file
            if not self.checker.load_excel_file(file_path):
                print("❌ Failed to load Excel file. Please check the file and try again.")
                return None
            
            # Generate comprehensive assessment
            print("\n📊 Performing comprehensive analysis...")
            assessment = self.checker.generate_comprehensive_report()
            
            # Store for potential export
            self.last_assessment = assessment
            
            # Display results
            self._display_results(assessment)
            
            return assessment
            
        except Exception as e:
            print(f"❌ Critical error during analysis: {str(e)}")
            print("This may be due to file corruption, unsupported Excel features, or system limitations.")
            return None
    
    def _display_results(self, assessment: AutomationAssessment) -> None:
        """
        Display assessment results in a formatted, professional manner.
        
        Args:
            assessment (AutomationAssessment): Assessment results to display
        """
        try:
            print("\n" + "=" * 80)
            print("📋 EXCEL AUTOMATION FEASIBILITY REPORT")
            print("=" * 80)
            
            # Header information
            print(f"File: {assessment.file_info.get('name', 'Unknown')}")
            print(f"Analysis Date: {assessment.analysis_timestamp}")
            print(f"File Size: {assessment.file_info.get('size_mb', 0):.2f} MB")
            print(f"Sheets: {assessment.file_info.get('sheets', 0)}")
            
            # Overall assessment
            print(f"\n🎯 OVERALL ASSESSMENT")
            print("-" * 50)
            print(f"Automation Score: {assessment.overall_score}/100")
            
            # Color-code the feasibility level
            if assessment.overall_score >= 80:
                print(f"Feasibility Level: ✅ {assessment.feasibility_level}")
            elif assessment.overall_score >= 50:
                print(f"Feasibility Level: ⚠️  {assessment.feasibility_level}")
            else:
                print(f"Feasibility Level: ❌ {assessment.feasibility_level}")
            
            print(f"Estimated Effort: {assessment.estimated_effort}")
            
            # Detailed scores
            if 'scores' in assessment.detailed_analysis:
                print(f"\n📈 DETAILED SCORES")
                print("-" * 50)
                scores = assessment.detailed_analysis['scores']
                print(f"• File Structure Score: {scores.get('structure_score', 0)}/100")
                print(f"• Formula Complexity Score: {scores.get('formula_difficulty_score', 0)}/100")
                print(f"• Automation Pattern Score: {scores.get('pattern_score', 0)}/100")
            
            # Technical analysis
            if 'formula_analysis' in assessment.detailed_analysis:
                print(f"\n🔧 TECHNICAL ANALYSIS")
                print("-" * 50)
                formula_info = assessment.detailed_analysis['formula_analysis']
                print(f"• Total Formulas: {formula_info.get('total_formulas', 0):,}")
                print(f"• Complex Formulas: {formula_info.get('complex_formulas', 0):,}")
                print(f"• Simple Formulas: {formula_info.get('simple_formulas', 0):,}")
                print(f"• Complexity Ratio: {formula_info.get('complexity_ratio', 0):.1%}")
                
                # Show top formula types if available
                formula_types = formula_info.get('formula_types', {})
                if formula_types:
                    print("• Most Used Functions:", ", ".join(f"{k}({v})" for k, v in sorted(formula_types.items(), key=lambda x: x[1], reverse=True)[:5]))
            
            # Pattern analysis
            if 'pattern_analysis' in assessment.detailed_analysis:
                print(f"\n🔄 BUSINESS PROCESS PATTERNS")
                print("-" * 50)
                pattern_info = assessment.detailed_analysis['pattern_analysis']
                
                detected_patterns = pattern_info.get('detected_patterns', [])
                if detected_patterns:
                    for pattern in detected_patterns:
                        print(f"• {pattern}")
                else:
                    print("• No clear automation patterns detected")
                
                business_indicators = pattern_info.get('business_indicators', [])
                if business_indicators:
                    print("\nBusiness Process Indicators:")
                    for indicator in business_indicators:
                        print(f"• {indicator}")
            
            # Opportunities
            if assessment.opportunities:
                print(f"\n✅ AUTOMATION OPPORTUNITIES")
                print("-" * 50)
                for i, opportunity in enumerate(assessment.opportunities, 1):
                    print(f"{i}. {opportunity}")
            
            # Red flags
            if assessment.red_flags:
                print(f"\n⚠️  RED FLAGS & CHALLENGES")
                print("-" * 50)
                for i, flag in enumerate(assessment.red_flags, 1):
                    print(f"{i}. {flag}")
            
            # Tool recommendations
            print(f"\n🛠️  RECOMMENDED AUTOMATION APPROACHES")
            print("-" * 50)
            if assessment.recommended_tools:
                for i, tool in enumerate(assessment.recommended_tools, 1):
                    print(f"{i}. {tool}")
            else:
                print("No specific tool recommendations available")
            
            # Next steps
            print(f"\n💡 RECOMMENDED NEXT STEPS")
            print("-" * 50)
            
            if assessment.overall_score >= 70:
                print("✅ PROCEED WITH AUTOMATION:")
                print("   → This file is an excellent candidate for automation")
                print("   → Start with the highest-rated tools above")
                print("   → Focus on the identified opportunities")
                print("   → Consider starting with a pilot implementation")
                
            elif assessment.overall_score >= 50:
                print("⚠️  PROCEED WITH PREPARATION:")
                print("   → This file can be automated with some preparation work")
                print("   → Address the identified red flags first")
                print("   → Consider restructuring the most problematic areas")
                print("   → Plan for longer development timeline")
                
            elif assessment.overall_score >= 30:
                print("🔄 RESTRUCTURE BEFORE AUTOMATION:")
                print("   → Significant preparation work required")
                print("   → Consider redesigning the most complex parts")
                print("   → Evaluate if manual process redesign is better")
                print("   → Use RPA tools if structure cannot be changed")
                
            else:
                print("❌ AUTOMATION NOT RECOMMENDED:")
                print("   → File requires major restructuring before automation")
                print("   → Consider complete process redesign")
                print("   → Manual optimization may be more cost-effective")
                print("   → Consult with process improvement specialists")
            
            # Quality indicators
            if 'quality_metrics' in assessment.detailed_analysis:
                quality = assessment.detailed_analysis['quality_metrics']
                completeness = quality.get('analysis_completeness', 0)
                if completeness < 100:
                    print(f"\n📊 Analysis Completeness: {completeness:.0f}%")
                    if completeness < 80:
                        print("⚠️  Some analysis components failed - results may be incomplete")
            
            print("\n" + "=" * 80)
            
        except Exception as e:
            print(f"❌ Error displaying results: {str(e)}")
    
    def run_interactive_session(self) -> None:
        """
        Run an interactive session for analyzing Excel files.
        This is the main entry point for user interaction.
        """
        try:
            print("\n" + "=" * 80)
            print("🤖 ENTERPRISE EXCEL AUTOMATION FEASIBILITY CHECKER")
            print("=" * 80)
            print("Analyze Excel files to determine automation potential")
            print("Designed for enterprise environments with security compliance")
            print("\nVersion: 1.0 | Secure Local Processing Only")
            print("=" * 80)
            
            # Check requirements first
            if not self.check_requirements():
                print("\n❌ Cannot proceed due to missing requirements.")
                return
            
            while True:
                try:
                    print(f"\n🔧 MAIN MENU")
                    print("-" * 30)
                    print("1. Analyze Excel file")
                    print("2. Export last report to file")
                    print("3. Show system information")
                    print("4. Exit")
                    
                    choice = input("\nEnter your choice (1-4): ").strip()
                    
                    if choice == '1':
                        self._handle_file_analysis()
                    elif choice == '2':
                        self._handle_report_export()
                    elif choice == '3':
                        self._show_system_info()
                    elif choice == '4':
                        print("\n👋 Thank you for using Excel Automation Feasibility Checker!")
                        print("Remember: All analysis was performed locally with no external data transmission.")
                        break
                    else:
                        print("❌ Invalid choice. Please enter 1, 2, 3, or 4.")
                        
                except KeyboardInterrupt:
                    print("\n\n🛑 Operation cancelled by user.")
                    print("Exiting safely...")
                    break
                except Exception as e:
                    print(f"❌ Error in main menu: {str(e)}")
                    continue
                    
        except Exception as e:
            print(f"❌ Critical error in interactive session: {str(e)}")
    
    def _handle_file_analysis(self) -> None:
        """Handle the file analysis workflow."""
        try:
            print(f"\n📁 FILE ANALYSIS")
            print("-" * 30)
            
            # Get file path from user
            while True:
                file_path = input("Enter the full path to your Excel file: ").strip()
                
                if not file_path:
                    print("❌ Please provide a valid file path.")
                    continue
                
                # Handle quoted paths
                file_path = file_path.strip('"\'')
                
                if os.path.exists(file_path):
                    break
                else:
                    print(f"❌ File not found: {file_path}")
                    retry = input("Would you like to try another path? (y/n): ").strip().lower()
                    if retry != 'y':
                        return
            
            # Perform analysis
            assessment = self.analyze_file(file_path)
            
            if assessment:
                # Ask if user wants to export report
                export_choice = input("\n💾 Would you like to export this report to a text file? (y/n): ").strip().lower()
                if export_choice == 'y':
                    try:
                        output_path = self.checker.export_report_to_text(assessment)
                        if output_path:
                            print(f"✅ Report exported to: {output_path}")
                        else:
                            print("❌ Failed to export report")
                    except Exception as e:
                        print(f"❌ Export failed: {str(e)}")
            
        except Exception as e:
            print(f"❌ Error in file analysis workflow: {str(e)}")
    
    def _handle_report_export(self) -> None:
        """Handle exporting the last analysis report."""
        try:
            if not self.last_assessment:
                print("❌ No analysis results available to export.")
                print("Please analyze a file first using option 1.")
                return
            
            print(f"\n💾 EXPORT REPORT")
            print("-" * 30)
            
            # Get custom output path if desired
            custom_path = input("Enter custom output path (or press Enter for auto-generated): ").strip()
            
            output_path = custom_path if custom_path else None
            
            try:
                exported_path = self.checker.export_report_to_text(self.last_assessment, output_path)
                if exported_path:
                    print(f"✅ Report successfully exported to: {exported_path}")
                else:
                    print("❌ Export failed")
            except Exception as e:
                print(f"❌ Export error: {str(e)}")
                
        except Exception as e:
            print(f"❌ Error in report export workflow: {str(e)}")
    
    def _show_system_info(self) -> None:
        """Display system and version information."""
        try:
            print(f"\n💻 SYSTEM INFORMATION")
            print("-" * 30)
            print(f"Python Version: {sys.version}")
            print(f"Operating System: {os.name}")
            print(f"Current Working Directory: {os.getcwd()}")
            
            # Library versions
            try:
                import pandas as pd
                print(f"Pandas Version: {pd.__version__}")
            except:
                print("Pandas Version: Not available")
            
            try:
                import openpyxl
                print(f"OpenPyXL Version: {openpyxl.__version__}")
            except:
                print("OpenPyXL Version: Not available")
            
            print(f"\n🔒 SECURITY COMPLIANCE")
            print("-" * 30)
            print("✅ No external network connections")
            print("✅ No data transmission outside local environment")
            print("✅ All processing performed locally")
            print("✅ Uses only standard enterprise-approved libraries")
            print("✅ No sensitive data stored or cached")
            
        except Exception as e:
            print(f"❌ Error displaying system info: {str(e)}")

def main():
    """
    Main function to run the Excel Automation Feasibility Checker.
    
    This function initializes the bot and starts the interactive session.
    It includes comprehensive error handling for enterprise environments.
    """
    try:
        # Initialize the bot
        bot = ExcelAutomationBot(verbose=True)
        
        # Run interactive session
        bot.run_interactive_session()
        
    except KeyboardInterrupt:
        print("\n\n🛑 Program interrupted by user. Exiting safely...")
    except Exception as e:
        print(f"\n❌ Critical system error: {str(e)}")
        print("Please contact your IT administrator if this error persists.")
    finally:
        # Cleanup (if needed)
        print("\n🔒 Session ended. No data retained or transmitted.")

def analyze_single_file(file_path: str, export_report: bool = True, verbose: bool = True) -> Optional[AutomationAssessment]:
    """
    Convenience function to analyze a single Excel file programmatically.
    
    This function is useful for integration with other systems or batch processing.
    
    Args:
        file_path (str): Path to the Excel file to analyze
        export_report (bool): Whether to automatically export a text report
        verbose (bool): Whether to show detailed progress information
        
    Returns:
        Optional[AutomationAssessment]: Assessment results or None if failed
    """
    try:
        # Create bot instance
        bot = ExcelAutomationBot(verbose=verbose)
        
        # Check requirements
        if not bot.check_requirements():
            return None
        
        # Analyze file
        assessment = bot.analyze_file(file_path)
        
        # Export report if requested and analysis succeeded
        if assessment and export_report:
            try:
                output_path = bot.checker.export_report_to_text(assessment)
                if output_path and verbose:
                    print(f"\n💾 Report automatically exported to: {output_path}")
            except Exception as e:
                if verbose:
                    print(f"⚠️  Report export failed: {str(e)}")
        
        return assessment
        
    except Exception as e:
        if verbose:
            print(f"❌ Error in single file analysis: {str(e)}")
        return None

# Enterprise usage examples and documentation
def show_usage_examples():
    """
    Display usage examples for different enterprise scenarios.
    """
    print("\n" + "=" * 80)
    print("📚 USAGE EXAMPLES FOR ENTERPRISE ENVIRONMENTS")
    print("=" * 80)
    
    print("\n1. INTERACTIVE MODE (Recommended for initial assessments):")
    print("   python excel_automation_checker.py")
    print("   # Follow the interactive prompts")
    
    print("\n2. PROGRAMMATIC MODE (For integration with other systems):")
    print("   from excel_automation_checker import analyze_single_file")
    print("   assessment = analyze_single_file('path/to/file.xlsx')")
    print("   if assessment:")
    print("       print(f'Score: {assessment.overall_score}')")
    
    print("\n3. BATCH PROCESSING (For multiple files):")
    print("   import os")
    print("   from excel_automation_checker import ExcelAutomationBot")
    print("   ")
    print("   bot = ExcelAutomationBot()")
    print("   for file in os.listdir('excel_files/'):")
    print("       if file.endswith(('.xlsx', '.xlsm')):")
    print("           assessment = bot.analyze_file(file)")
    
    print("\n4. ENTERPRISE INTEGRATION:")
    print("   # Can be integrated with:")
    print("   • SharePoint document libraries")
    print("   • File system monitoring")
    print("   • Business process management systems")
    print("   • Project planning tools")

if __name__ == "__main__":
    """
    Entry point for the Excel Automation Feasibility Checker.
    
    This script can be run in multiple ways:
    1. Interactive mode (default)
    2. Command line with file path
    3. Imported as a module for programmatic use
    """
    
    # Check if file path provided as command line argument
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        print(f"\n🚀 Analyzing file from command line: {file_path}")
        
        # Check for additional flags
        verbose = '--quiet' not in sys.argv
        export_report = '--no-export' not in sys.argv
        
        # Analyze single file
        assessment = analyze_single_file(file_path, export_report, verbose)
        
        if assessment:
            print(f"\n✅ Analysis complete. Overall score: {assessment.overall_score}/100")
            print(f"Feasibility: {assessment.feasibility_level}")
        else:
            print("\n❌ Analysis failed")
            sys.exit(1)
    else:
        # Run interactive mode
        main()